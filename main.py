from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd
import bs4
import csv
from openpyxl import Workbook
import os

app = Flask(__name__)
upload_path = 'convertr/'

@app.route("/")
def home():
    return render_template("index.html")


@app.route("/convert/<var>/<err>")
def convert(var, err):
    return render_template("convert.html", variable = var, err = err)


@app.route("/download", methods=['POST'])
def download():
    if request.method == 'POST':
        name = request.form.get('name')
        html = request.form.get('html')
        try:
            f = open(upload_path+name+".html", "w")
            f.write(html)
            f.close()
            return send_file(name+'.html', as_attachment=True)
        except:
            return render_template('download.html', html = "Due to some error, file cannot be downloaded!", err = "1", name = "0")
        finally:
            if(os.path.isfile(upload_path+name+'.html')):
                os.remove(upload_path+name+'.html')


def x2h(ename, name):
    try:
        df = pd.read_excel(ename)
        if df.empty:
            return render_template('download.html', html = "The file is empty, No HTML generated!", err = "1", name="0")
        html = df.to_html(index=False)
        html = html.replace("NaN","")
        return render_template('download.html', html = html, err = "0", name = name)
    except:
        return render_template('download.html', html = "Due to some error, HTML cannot be generated!", err = "1", name = "0")
    finally:
        if(os.path.isfile(ename)):
            os.remove(ename)


def c2h(fname, name):
    try:
        df = pd.read_csv(fname)
        html = df.to_html(index=False)
        return render_template('download.html', html = html, err = "0", name = name)
    except:
        return render_template('download.html', html = "CSV is blank or unable to convert!", err = "1", name="0")
    finally:
        if os.path.isfile(fname):
            os.remove(fname)


def h2x(hname, name):
    try:
        with open(hname, 'r', encoding='utf-8') as html_file:
            html_content = html_file.read()
        soup = bs4.BeautifulSoup(html_content, 'html.parser')
        tables = soup.find_all('table')
        if not tables:
            return render_template('download.html', html = "There are no tables or unable to detect!", err = "1", name="0")
        else:
            workbook = Workbook()
            for table_index, table in enumerate(tables, start=1):
                worksheet = workbook.create_sheet(title=f'Table_{table_index}')
                for row_index, row in enumerate(table.find_all('tr'), start=1):
                    for col_index, cell in enumerate(row.find_all(['td', 'th']), start=1):
                        worksheet.cell(row=row_index, column=col_index, value=cell.get_text(strip=True))

            workbook.remove(workbook.active)
            workbook.save(upload_path+name+'.xlsx')
        return send_file(name+'.xlsx', as_attachment=True)
    except:
        return render_template('download.html', html = "Due to some error, File cannot be converted!", err = "1", name="0")
    finally:
        if os.path.isfile(hname):
            os.remove(hname)
        if os.path.isfile(upload_path+name+'.xlsx'):
            os.remove(upload_path+name+'.xlsx')


def h2c(hname, name):
    try:
        with open(hname, 'r', encoding='utf-8') as html_file:
            html_content = html_file.read()
        soup = bs4.BeautifulSoup(html_content, 'html.parser')
        tables = soup.find_all('table')
        if not tables:
            return render_template('download.html', html = "There are no tables or unable to detect!", err = "1", name="0")
        else:
            csv_filename = upload_path+name+'.csv'
            with open(csv_filename, 'a', newline='', encoding='utf-8') as csv_file:
                csv_writer = csv.writer(csv_file)

                for table_index, table in enumerate(tables, start=1):
                    if table_index > 1:
                        csv_writer.writerow([])
                        csv_writer.writerow([f"Table-{table_index}"])
                        csv_writer.writerow([])

                    for row in table.find_all('tr'):
                        csv_writer.writerow(cell.get_text(strip=True) for cell in row.find_all(['td', 'th']))

        return send_file(name+'.csv', as_attachment=True)
    except:
        return render_template('download.html', html = "Due to some error, File cannot be converted!", err = "1", name = "0")
    finally:
        if os.path.isfile(hname):
            os.remove(hname)
        if os.path.isfile(upload_path+name+'.csv'):
            os.remove(upload_path+name+'.csv')


@app.route('/success/<var>', methods = ['POST'])
def success(var):
    if request.method == 'POST':
        f = request.files['file']
        try:
            f.save(upload_path+f.filename)
        except:
            return redirect(url_for('convert', var = var, err = "File cannot be fetched. Try again!"))

        if var == 'x2h':
            name = f.filename.replace(".xlsx", "")
            return x2h(upload_path+f.filename, name)
        elif var == 'c2h':
            name = f.filename.replace(".csv", "")
            return c2h(upload_path+f.filename, name)
        elif var == 'h2x':
            name = f.filename.replace(".html", "")
            return h2x(upload_path+f.filename, name)
        elif var == 'h2c':
            name = f.filename.replace(".html", "")
            return h2c(upload_path+f.filename, name)
        return redirect(url_for('home'))